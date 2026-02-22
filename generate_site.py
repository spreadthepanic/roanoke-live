#!/usr/bin/env python3
"""
generate_site.py
Reads Upcoming_Entertainment.xlsx and generates index.html for GitHub Pages.

Usage:
    python generate_site.py
    python generate_site.py --file path/to/Upcoming_Entertainment.xlsx
    python generate_site.py --file Upcoming_Entertainment.xlsx --out index.html
"""

import argparse
import json
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

# Map venue names to their event calendar URLs.
# Add or update entries here as needed.
VENUE_URLS = {
    "Martin's":         "https://www.martinsdowntown.com/live-music",
    "Harvester":        "https://www.harvester-music.com/",
    "Grandin Theater":  "https://www.grandintheatre.com/shows?tag=Live%20Music",
    "Berglund":         "https://berglundcenter.live/events",
    "Jefferson Center": "https://www.jeffcenter.org/events",
    "5 Points":         "https://www.5pointsmusic.com/events/live-events/",
    "Parkway":          "https://parkwaybrewing.com/the-tasting-room/calendar/",
    "Salem CC":         "https://www.salemciviccenter.com/calendar.aspx?CID=26",
    "Spot on Kirk":     "https://www.thespotonkirk.org/shows",
}


def load_events(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb["Music"]
    events = []
    today = date.today()
    for row in ws.iter_rows(min_row=3, values_only=True):
        dt, band, dave, karla, daisy, venue = row[1], row[2], row[3], row[4], row[5], row[6]
        if not (dt and band and venue):
            continue
        if isinstance(dt, datetime):
            dt = dt.date()
        if dt < today:
            continue
        events.append({
            "date":        dt.strftime("%Y-%m-%d"),
            "dateDisplay": dt.strftime("%a, %b ") + str(dt.day),
            "band":        band,
            "venue":       venue,
            "venueUrl":    VENUE_URLS.get(venue, ""),
            "dave":        dave == "Yes",
            "karla":       karla == "Yes",
            "daisy":       daisy == "Yes",
        })
    return events


def make_event_card(e):
    key = f"{e['date']}|{e['band']}"
    safe_id = "".join(c if c.isalnum() else "_" for c in key)
    venue_html = (
        f'<a href="{e["venueUrl"]}" target="_blank" rel="noopener">📍 {e["venue"]}</a>'
        if e["venueUrl"] else f'📍 {e["venue"]}'
    )
    has_any_ticket = e["dave"] or e["karla"] or e["daisy"]
    ticket_badge = '<span class="ticket-badge">tickets secured</span>' if has_any_ticket else ""
    return f"""  <div class="event" id="card_{safe_id}">
    <div class="event-date">{e["dateDisplay"]}</div>
    <div class="event-info">
      <div class="event-band"><strong>{e["band"]}</strong>{ticket_badge}</div>
      <div class="event-meta">{venue_html}<span class="event-meta-extra"></span></div>
    </div>
    <div class="event-action"></div>
  </div>"""


def make_venue_button(venue):
    safe = venue.replace("'", "\\'")
    return f'<button class="small venue-btn" data-venue="{venue}" onclick="setVenue(\'{safe}\')">{venue}</button>'


HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Roanoke Live</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    min-height: 100vh;
    background: #0b0f1a;
    color: #dde3ee;
    font-family: 'Palatino Linotype', Palatino, Georgia, serif;
  }}
  #header {{
    background: linear-gradient(160deg, #0d1b3e 0%, #0b2444 60%, #0a1628 100%);
    padding: 36px 20px 28px;
    border-bottom: 1px solid #1c2f50;
  }}
  #header-inner {{ max-width: 740px; margin: 0 auto; }}
  h1 {{ font-size: 32px; font-weight: normal; color: #e8f0fe; letter-spacing: 0.03em; margin-bottom: 4px; }}
  .subtitle {{ color: #5a7a9a; font-size: 13px; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 22px; }}
  .toggle-row {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  .toggle-label {{ font-size: 13px; color: #5a7a9a; }}
  button {{
    font-family: inherit; cursor: pointer; transition: all 0.2s;
    border-radius: 20px; padding: 6px 18px; font-size: 13px; border: 1px solid #1c2f50;
    background: #111d35; color: #7a9ab5;
  }}
  button.active {{ background: #2dd4bf; color: #0b0f1a; border-color: #2dd4bf; font-weight: bold; }}
  button.small {{ border-radius: 16px; padding: 4px 13px; font-size: 12px; }}
  #venue-bar {{
    background: #0c1220; border-bottom: 1px solid #1a2740;
    overflow-x: auto; white-space: nowrap;
  }}
  #venue-inner {{ padding: 10px 20px; max-width: 740px; margin: 0 auto; display: flex; gap: 6px; }}
  #stats-bar {{
    background: #090d18; padding: 8px 20px; border-bottom: 1px solid #131e30;
    font-size: 12px; color: #3a5a7a;
  }}
  #stats-inner {{ max-width: 740px; margin: 0 auto; display: flex; gap: 16px; flex-wrap: wrap; }}
  .stat-yes {{ color: #2dd4bf; }}
  #events {{ max-width: 740px; margin: 0 auto; padding: 16px 20px; }}
  .event {{
    border-radius: 8px; padding: 12px 16px; margin-bottom: 8px;
    display: flex; gap: 14px; align-items: center; flex-wrap: wrap;
    border: 1px solid #141e30; border-left: 3px solid #1c2f50;
    background: #0f1525; transition: border-color 0.25s;
  }}
  .event.has-ticket {{ background: #0d1e33; border-color: #1a3455; border-left-color: #2dd4bf; }}
  .event.rsvp-yes   {{ border-left-color: #f59e0b; }}
  .event.rsvp-maybe {{ border-left-color: #f59e0b88; }}
  .event-date {{ min-width: 84px; font-size: 12px; color: #4a6a8a; line-height: 1.4; flex-shrink: 0; }}
  .event-info {{ flex: 1; min-width: 160px; }}
  .event-band {{ font-size: 15px; color: #dde3ee; margin-bottom: 4px; line-height: 1.3; }}
  .ticket-badge {{
    font-size: 10px; background: #2dd4bf20; color: #2dd4bf;
    padding: 2px 8px; border-radius: 10px; letter-spacing: 0.08em;
    text-transform: uppercase; vertical-align: middle; margin-left: 8px;
  }}
  .event-meta {{ font-size: 13px; color: #5a7a9a; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  .event-meta a {{ color: #5a7a9a; text-decoration: none; }}
  .event-meta a:hover {{ color: #2dd4bf; }}
  .confirmed {{ color: #2dd4bf; }}
  .others    {{ color: #f59e0b; }}
  .rsvp-btn {{
    border-radius: 5px; padding: 5px 10px; font-size: 12px;
    font-family: inherit; cursor: pointer; white-space: nowrap;
    background: transparent; color: #3a5a7a; border: 1px solid #1c2f50; transition: all 0.15s;
  }}
  .rsvp-btn.sel-yes   {{ background: #2dd4bf33; color: #2dd4bf; border-color: #2dd4bf; }}
  .rsvp-btn.sel-maybe {{ background: #f59e0b33; color: #f59e0b; border-color: #f59e0b; }}
  .rsvp-btn.sel-no    {{ background: #4b556333; color: #9ca3af; border-color: #4b5563; }}
  .going-badge {{
    font-size: 12px; color: #2dd4bf; background: #2dd4bf15;
    border: 1px solid #2dd4bf33; border-radius: 6px; padding: 6px 12px;
    white-space: nowrap; flex-shrink: 0;
  }}
  .rsvp-area {{ display: flex; gap: 5px; flex-wrap: wrap; flex-shrink: 0; }}
  #empty {{ text-align: center; color: #2a3a5a; padding: 60px 0; font-size: 15px; display: none; }}
  footer {{ text-align: center; padding: 20px; color: #1c2f50; font-size: 12px; }}
</style>
</head>
<body>

<div id="header">
  <div id="header-inner">
    <h1>&#127925; Roanoke Live</h1>
    <p class="subtitle">Upcoming shows &middot; Roanoke area</p>
    <div class="toggle-row">
      <span class="toggle-label">Viewing as:</span>
      <button class="person-btn active" data-person="karla" onclick="setPerson('karla')">&#127925; Karla</button>
      <button class="person-btn"        data-person="dave"  onclick="setPerson('dave')">&#127928; Dave</button>
      <button class="person-btn"        data-person="daisy" onclick="setPerson('daisy')">&#127804; Daisy</button>
      <span style="flex:1"></span>
      <button id="tickets-btn" onclick="toggleTickets()">&#127903; Tickets only</button>
    </div>
  </div>
</div>

<div id="venue-bar">
  <div id="venue-inner">
    <button class="small venue-btn active" data-venue="All Venues" onclick="setVenue('All Venues')">All Venues</button>
    {venue_buttons}
  </div>
</div>

<div id="stats-bar">
  <div id="stats-inner">
    <span id="stat-shown"></span>
    <span>&middot;</span>
    <span id="stat-tickets"></span>
    <span id="stat-rsvp" class="stat-yes" style="display:none"></span>
  </div>
</div>

<div id="events">
{event_cards}
  <div id="empty">No shows match this filter.</div>
</div>

<footer>Curated by Dave &middot; Roanoke live music &middot; Updated {updated}</footer>

<script>
const EVENTS = {events_json};

let currentPerson = 'karla';
let currentVenue  = 'All Venues';
let ticketsOnly   = false;
let rsvps         = {{}};

function loadRsvps() {{
  try {{ rsvps = JSON.parse(localStorage.getItem('rsvps-' + currentPerson) || '{{}}'); }}
  catch {{ rsvps = {{}}; }}
}}

function saveRsvp(key, val) {{
  if (rsvps[key] === val) delete rsvps[key]; else rsvps[key] = val;
  try {{ localStorage.setItem('rsvps-' + currentPerson, JSON.stringify(rsvps)); }} catch {{}}
  render();
}}

function setPerson(p) {{
  currentPerson = p;
  document.querySelectorAll('.person-btn').forEach(b =>
    b.classList.toggle('active', b.dataset.person === p));
  loadRsvps();
  render();
}}

function setVenue(v) {{
  currentVenue = v;
  document.querySelectorAll('.venue-btn').forEach(b =>
    b.classList.toggle('active', b.dataset.venue === v));
  render();
}}

function toggleTickets() {{
  ticketsOnly = !ticketsOnly;
  document.getElementById('tickets-btn').classList.toggle('active', ticketsOnly);
  render();
}}

function render() {{
  const otherNames = {{ dave:'Dave', karla:'Karla', daisy:'Daisy' }};
  const others = ['dave','karla','daisy'].filter(p => p !== currentPerson);
  const totalTickets = EVENTS.filter(e => e.dave || e.karla || e.daisy).length;
  let shown = 0, rsvpYes = 0;

  EVENTS.forEach(e => {{
    const key = e.date + '|' + e.band;
    const safeId = key.replace(/[^a-z0-9]/gi, '_');
    const card = document.getElementById('card_' + safeId);
    if (!card) return;

    const visible = (currentVenue === 'All Venues' || e.venue === currentVenue)
                 && (!ticketsOnly || e[currentPerson]);
    card.style.display = visible ? '' : 'none';
    if (!visible) return;
    shown++;

    const rsvp  = rsvps[key] || null;
    const iHave = e[currentPerson];
    if (rsvp === 'yes') rsvpYes++;

    const othersWithTicket = others.filter(p => e[p]).map(p => otherNames[p]);

    // Border class
    card.className = 'event' + (iHave ? ' has-ticket'
                              : rsvp === 'yes' ? ' rsvp-yes'
                              : rsvp === 'maybe' ? ' rsvp-maybe' : '');

    // Ticket/RSVP status line
    const metaExtra = card.querySelector('.event-meta-extra');
    metaExtra.innerHTML = iHave
      ? '<span class="confirmed">&#10003; Your ticket confirmed</span>'
      : othersWithTicket.length
        ? '<span class="others">' + othersWithTicket.join(' &amp; ')
          + (othersWithTicket.length === 1 ? ' has' : ' have') + ' a ticket</span>'
        : '';

    // Action area
    const action = card.querySelector('.event-action');
    if (iHave) {{
      action.innerHTML = '<div class="going-badge">&#127903; You\'re going!</div>';
    }} else {{
      const btns = [['yes',"I'm in! 🎶"],['maybe','Maybe 🤔'],['no',"Can't go"]]
        .map(([v, l]) =>
          `<button class="rsvp-btn ${{rsvp === v ? 'sel-' + v : ''}}"
            onclick="saveRsvp('${{key.replace(/'/g, "\\'")}}','${{v}}')">${{l}}</button>`)
        .join('');
      action.innerHTML = '<div class="rsvp-area">' + btns + '</div>';
    }}
  }});

  document.getElementById('stat-shown').textContent =
    shown + ' show' + (shown === 1 ? '' : 's') + ' shown';
  document.getElementById('stat-tickets').textContent =
    totalTickets + ' with tickets secured';
  const rsvpEl = document.getElementById('stat-rsvp');
  if (rsvpYes > 0) {{
    rsvpEl.innerHTML = '&#10003; ' + rsvpYes + ' you\'re attending';
    rsvpEl.style.display = '';
  }} else {{
    rsvpEl.style.display = 'none';
  }}
  document.getElementById('empty').style.display = shown === 0 ? '' : 'none';
}}

loadRsvps();
render();
</script>
</body>
</html>
"""


def generate(xlsx_path, output_path):
    events = load_events(xlsx_path)
    venues = sorted(set(e["venue"] for e in events))
    venue_buttons = "\n    ".join(make_venue_button(v) for v in venues)
    event_cards   = "\n".join(make_event_card(e) for e in events)
    html = HTML_TEMPLATE.format(
        venue_buttons=venue_buttons,
        event_cards=event_cards,
        events_json=json.dumps(events, ensure_ascii=False),
        updated=datetime.today().strftime("%B ") + str(datetime.today().day) + datetime.today().strftime(", %Y"),
    )
    Path(output_path).write_text(html, encoding="utf-8")
    print(f"Generated {output_path}  ({len(events)} events, {len(venues)} venues)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Roanoke Live site from Excel.")
    parser.add_argument("--file", default="Upcoming_Entertainment.xlsx")
    parser.add_argument("--out",  default="index.html")
    args = parser.parse_args()
    generate(args.file, args.out)
